from transformers import RobertaForSequenceClassification, AutoTokenizer, AutoModelForTokenClassification, pipeline
import torch
import nltk
import docx2txt
import pandas as pd
import os
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.colors import WHITE
import gradio as gr

nltk.download('punkt')

# Load the model and tokenizer
senti_model = RobertaForSequenceClassification.from_pretrained("wonrax/phobert-base-vietnamese-sentiment")
senti_tokenizer = AutoTokenizer.from_pretrained("wonrax/phobert-base-vietnamese-sentiment", use_fast=False)

# Load the model and tokenizer segmentation
seg_tokenizer = AutoTokenizer.from_pretrained("NlpHUST/vi-word-segmentation")
seg_model = AutoModelForTokenClassification.from_pretrained("NlpHUST/vi-word-segmentation")
nlp = pipeline("token-classification", model=seg_model, tokenizer=seg_tokenizer)


# Word segmented
def segmentation(sentences):
    segmented_sentences = []
    for sentence in sentences:
        ner_results = nlp(sentence)
        sentence_tok = ""
        for e in ner_results:
            if "##" in e["word"]:
                sentence_tok = sentence_tok + e["word"].replace("##", "")
            elif e["entity"] == "I":
                sentence_tok = sentence_tok + "_" + e["word"]
            else:
                sentence_tok = sentence_tok + " " + e["word"]
        segmented_sentences.append(sentence_tok.strip())
    return segmented_sentences


# File read
def read_file(docx):
    try:
        text = docx2txt.process(docx)
        lines = text.split('\n')
        lines = [line.strip() for line in lines]
        lines = [line for line in lines if line]
        return lines  # add this line
    except Exception as e:
        print(f"Error reading file: {e}")


# Define a function to analyze the sentiment of a text
def analyze(sentence):
    input_ids = torch.tensor([senti_tokenizer.encode(sentence)])
    with torch.no_grad():
        out = senti_model(input_ids)
        results = out.logits.softmax(dim=-1).tolist()
        return results[0]


def file_analysis(docx):
    # Read the file and segment the sentences
    sentences = read_file(docx)
    segmented_sentences = segmentation(sentences)
    
    # Analyze the sentiment of each sentence
    results = []
    for sentence in segmented_sentences:
        results.append(analyze(sentence))

    return results


def generate_pie_chart(df):
    # Calculate the average scores
    neg_avg = df['Negative'].mean()
    pos_avg = df['Positive'].mean()
    neu_avg = df['Neutral'].mean()

    # Create a new DataFrame with the average scores
    avg_df = pd.DataFrame({'Sentiment': ['Negative', 'Positive', 'Neutral'],
                           'Score': [neg_avg, pos_avg, neu_avg]})

    # Set custom colors for the pie chart
    colors = ['#BDBDBD', '#9ACD32', '#87CEFA']

    # Create a pie chart showing the average scores
    plt.pie(avg_df['Score'], labels=avg_df['Sentiment'], colors=colors, autopct='%1.1f%%')
    plt.title('Average Scores by Sentiment')

    # Save the pie chart as an image file in the static folder
    pie_chart_name = 'pie_chart.png'
    plt.savefig(pie_chart_name)
    plt.close()

    return pie_chart_name


def generate_excel_file(df):
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add column headers to the worksheet
    headers = ['Negative', 'Positive', 'Neutral', 'Text']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Set up cell formatting for each sentiment
    fill_dict = {
        'Negative': PatternFill(start_color='BDBDBD', end_color='BDBDBD', fill_type='solid'),
        'Positive': PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid'),
        'Neutral': PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
    }

    # Loop through each row of the input DataFrame and write data to the worksheet
    for row_num, row_data in df.iterrows():
        # Calculate the highest score and corresponding sentiment for this row
        sentiment_cols = ['Negative', 'Positive', 'Neutral']
        scores = [row_data[col] for col in sentiment_cols]
        max_score = max(scores)
        max_index = scores.index(max_score)
        sentiment = sentiment_cols[max_index]

        # Write the data to the worksheet
        for col_num, col_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num + 2, column=col_num)
            cell.value = col_data
            if col_num in [1, 2, 3]:
                if col_data == max_score:
                    cell.fill = fill_dict[sentiment]
            if col_num == 4:
                fill = fill_dict[sentiment]
                font_color = WHITE if fill.start_color.rgb == 'BDBDBD' else Color('000000')
                cell.fill = fill
                cell.font = Font(color=font_color)
                if col_data == max_score:
                    cell.fill = fill_dict[sentiment]

    # Save the workbook
    excel_file_path = 'result.xlsx'
    wb.save(excel_file_path)

    return excel_file_path


def process_file(docx):
    # Perform analysis on the file
    results = file_analysis(docx)

    # Create a DataFrame from the results
    df = pd.DataFrame(results, columns=['Negative', 'Positive', 'Neutral'])
    df['Text'] = read_file(docx)

    # Generate the pie chart and excel file
    pie_chart_name = generate_pie_chart(df)
    excel_file_path = generate_excel_file(df)

    return pie_chart_name, excel_file_path

def analyze_file(file, sentence):
    if file and sentence:
        # Both file and sentence inputs are provided
        # Process the uploaded file and generate the output files
        pie_chart_name, excel_file_path = process_file(file.name)

        # Analyze the sentiment of the input sentence
        segmented_sentence = segmentation([sentence])
        results = analyze(segmented_sentence[0])

        # Get the label names
        label_names = ['Negative', 'Positive', 'Neutral']

        # Create the output text with labels and scores
        output_text = ""
        for label, score in zip(label_names, results):
            score_formatted = "{:.2f}".format(score)
            output_text += f"{label}: {score_formatted}\n"

        return excel_file_path, pie_chart_name, output_text
    
    elif sentence:
        # Only sentence input is provided
        # Analyze the sentiment of the input sentence
        segmented_sentence = segmentation([sentence])
        results = analyze(segmented_sentence[0])

        # Get the label names
        label_names = ['Negative', 'Positive', 'Neutral']

        # Create the output text with labels and scores
        output_text = ""
        for label, score in zip(label_names, results):
            score_formatted = "{:.2f}".format(score)
            output_text += f"{label}: {score_formatted}\n"

        return None, None, output_text
    elif file:
        # Only file input is provided
        # Process the uploaded file and generate the output files
        pie_chart_name, excel_file_path = process_file(file.name)

        # Return the file paths for the pie chart and excel file
        return excel_file_path, pie_chart_name, None

inputs = [
    gr.File(label="Chọn Tệp Bạn Muốn Phân Tích"),
    gr.Textbox(label="Nhập Văn Bản")
]
outputs = [
    gr.File(label="Kết Quả Phân Tích Excel"),
    gr.Image(type="filepath", label="Thông Số Phân Tích"),
    gr.Textbox(label="Kết Quả Phân Tích")
]

interface = gr.Interface(
    fn=analyze_file,
    inputs=inputs,
    outputs=outputs,
    title="Sentiment Analysis",
    allow_flagging="never"  # Disable flag button
)


if __name__ == "__main__":
    interface.launch(share=True , debug=True)
